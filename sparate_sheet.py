import openpyxl as xl
wb = xl.load_workbook("transactions.xlsx")
sheet = wb['Sheet1']
cell= sheet['a1']
cell= sheet.cell(1,1)
print(cell.value)
print(sheet.max_row)
for row in range(2,sheet.max_row+1):
    cell=sheet.cell(row,3)
    corr_prize = cell.value*0.9
    corr_price_cell=sheet.cell(row,4)
    corr_price_cell.value=corr_prize

wb.save('trans2.xlsx')
